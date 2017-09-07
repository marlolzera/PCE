# -*- coding: utf-8 -*-
from __future__ import unicode_literals

from openpyxl import Workbook
from django.contrib.admin.utils import quote
from django.contrib.auth.decorators import login_required

import json
import datetime

from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from django.db.models import F
from django.http import HttpResponse

from interface.forms import UsuarioForm, UsuarioInfoForm
from interface.functions import render_to_pdf
from interface.models import UsuarioInfo
from principal.models import *
from principal.forms import ProdutoForm, CategoriaForm, FornecedorForm, EntradaProdutoForm, SaidaProdutoForm
from django.shortcuts import render, redirect
from django.shortcuts import render
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.models import User
from django.utils import timezone


@login_required(login_url='/login')
def home(request):
    # queryset = Produto.objects.values('categoria__nome', 'nome', 'quantidade').filter(quantidade=150)

    parametros_filtro = {

    }

    queryset = EntradaProduto.objects.annotate(
        lote_movimento=F('lote'),
        nome_produto=F('produto__nome'),
        nome_categoria=F('produto__categoria__nome'),
        fornecedor_usuario=F('fornecedor__nome_fantasia'),
        quantidade_mov=F('quantidade'),
        data_mov=F('data_cadastro')).extra(select={'tipo_mov': 1}).values(
        'lote_movimento',
        'nome_produto',
        'nome_categoria',
        'fornecedor_usuario',
        'quantidade_mov',
        'data_mov',
        'tipo_mov').filter(**parametros_filtro)

    return render(request, 'principal/index.html', locals())


@login_required(login_url='/login')
def atualiza_login(request, id):
    usuario = User.objects.get(id=id)
    usuario_info = UsuarioInfo.objects.get(user_id=id)

    usuario_form = UsuarioForm(request.POST or None, instance=usuario)
    usuarioinfo_form = UsuarioInfoForm(request.POST or None, instance=usuario_info)

    if request.method == 'POST':
        if usuario_form.is_valid() and usuarioinfo_form.is_valid():
            usuario = usuario_form.save()
            usuario.set_password(usuario.password)
            usuario.save()

            usuarioinfo = usuarioinfo_form.save(commit=False)
            usuarioinfo.user = usuario
            usuarioinfo.save()

            # desloga e loga o usuário
            username = request.POST['username']
            password = request.POST['password']
            logout(request)
            login(request, authenticate(username=username, password=password))

            return redirect('/home/')

    return render(request, 'interface/nova_conta.html',
                  {'usuario_form': usuario_form,
                   'usuarioinfo_form': usuarioinfo_form})


# ////////////////////////////////////////// Categoria /////////////////////////////////////////////


@login_required(login_url='/login')
def lista_categoria(request):
    lista_categoria = Categoria.objects.all().order_by('-id')
    return render(request, 'principal/lista_categoria.html', locals())


@login_required(login_url='/login')
def nova_categoria(request):
    form = CategoriaForm(request.POST or None)

    if request.method == 'POST':
        if form.is_valid():
            form.save()
            return redirect('/lista_categoria/')

    return render(request, 'principal/nova_categoria.html', {'form': form})


@login_required(login_url='/login')
def atualiza_categoria(request, id):
    categoria = Categoria.objects.get(id=id)
    form = CategoriaForm(request.POST or None, instance=categoria)

    if request.method == 'POST':
        if form.is_valid():
            form.save()
            return redirect('/lista_categoria/')

    return render(request, 'principal/nova_categoria.html', {'form': form})


@login_required(login_url='/login')
def deleta_categoria(request, id):
    categoria = Categoria.objects.get(id=id)
    categoria.delete()
    return redirect('/lista_categoria/')


# ////////////////////////////////////////// Produto ///////////////////////////////////////////////


@login_required(login_url='/login')
def novo_produto(request):
    form = ProdutoForm(request.POST or None)

    if request.method == 'POST':
        if form.is_valid():
            form.save()
            return redirect('/lista_produto/')

    return render(request, 'principal/novo_produto.html', {'form': form})


@login_required(login_url='/login')
def novo_produto_dialog(request):
    return render(request, 'principal/novo_produto_dialog.html')


@login_required(login_url='/login')
def lista_produto(request):
    lista_produto = Produto.objects.select_related().order_by('-id')

    return render(request, 'principal/lista_produto.html', locals())


@login_required(login_url='/login')
def atualiza_produto(request, id):
    produto = Produto.objects.get(id=id)
    form = ProdutoForm(request.POST or None, instance=produto)

    if request.method == 'POST':
        if form.is_valid():
            form.save()
            return redirect('/lista_produto/')

    return render(request, 'principal/novo_produto.html', {'form': form})


@login_required(login_url='/login')
def deleta_produto(request, id):
    produto = Produto.objects.get(id=id)
    produto.delete()
    return redirect('/lista_produto/')


# ////////////////////////////////////////// Fornecedor ///////////////////////////////////////////////


@login_required(login_url='/login')
def novo_fornecedor(request):
    form = FornecedorForm(request.POST or None)

    if request.method == 'POST':
        if form.is_valid():
            form.save()
            return redirect('/lista_fornecedor/')

    return render(request, 'principal/novo_fornecedor.html', {'form': form})


@login_required(login_url='/login')
def lista_fornecedor(request):
    lista_fornecedor = Fornecedor.objects.all().order_by('-id')

    return render(request, 'principal/lista_fornecedor.html', locals())


@login_required(login_url='/login')
def atualiza_fornecedor(request, id):
    fornecedor = Fornecedor.objects.get(id=id)
    form = FornecedorForm(request.POST or None, instance=fornecedor)

    if request.method == 'POST':
        if form.is_valid():
            form.save()
            return redirect('/lista_fornecedor/')

    return render(request, 'principal/novo_fornecedor.html', {'form': form})


@login_required(login_url='/login')
def deleta_fornecedor(request, id):
    fornecedor = Fornecedor.objects.get(id=id)
    fornecedor.delete()
    return redirect('/lista_fornecedor/')


# //////////////////////////////////// Entrada //////////////////////////////////////


@login_required(login_url='/login')
def entrada_produto(request):
    inicial = {}
    if request.POST:
        inicial.update({
            'categoria': request.POST.get("categoria")
        })

    form = EntradaProdutoForm(request.POST or None, initial=inicial)

    if request.method == 'POST':
        if form.is_valid():
            entrada_produto = form.save(commit=False)
            entrada_produto.quantidade = entrada_produto.quantidade_inicial
            entrada_produto.usuario = request.user
            entrada_produto.save()

            return redirect('/lista_lote/')

    return render(request, 'principal/entrada_produto.html', {'form': form})


@login_required(login_url='/login')
def lista_lote(request):
    # lista_lote = EntradaProduto.objects.select_related().order_by('-id')

    lista_lote = EntradaProduto.objects.annotate(
        nome_produto=F('produto__nome'),
        nome_categoria=F('produto__categoria__nome'),
        numero_lote=F('lote'),
        qtde_inicial=F('quantidade_inicial'),
        qtde_atual=F('quantidade'),
        dt_cadastro=F('data_cadastro'),
        dt_fabricacao=F('data_fabricacao'),
        dt_validade=F('data_validade')
    ).values(
        'nome_produto',
        'nome_categoria',
        'numero_lote',
        'qtde_inicial',
        'qtde_atual',
        'dt_cadastro',
        'dt_fabricacao',
        'dt_validade').order_by('produto__nome')

    return render(request, 'principal/lista_lote.html', locals())


# //////////////////////////////////// Saída //////////////////////////////////////


@login_required(login_url='/login')
def saida_produto(request):
    inicial = {}
    if request.POST:
        inicial.update({
            'categoria': request.POST.get("categoria"),
            'produto': request.POST.get("produto")
        })

    form = SaidaProdutoForm(request.POST or None, initial=inicial)

    if request.method == 'POST':
        if form.is_valid():
            saida_produto = form.save(commit=False)
            saida_produto.usuario = request.user
            saida_produto.save()

            return redirect('/lista_movimento/')

    return render(request, 'principal/saida_produto.html', {'form': form})


@login_required(login_url='/login')
def get_produto(request):
    if request.method != 'GET' or request.is_ajax() is False:
        pass
    else:
        categoria_id = request.GET.get('option', None)

        if categoria_id == '':
            categoria_id = '0'

        out = []

        try:
            produtos = Produto.objects.filter(categoria_id=categoria_id).order_by('nome')
        except Produto.DoesNotExist:
            produtos = None

        if produtos:
            out.append(u'{"id":0, "nome":"--- Selecione ---"}')
            for p in produtos:
                out.append(u'{"id":%s, "nome":"%s"}' % (p.id, p.nome))

        return HttpResponse(json.dumps(out), content_type='application/javascript')


@login_required(login_url='/login')
def get_lote(request):
    if request.method != 'GET' or request.is_ajax() is False:
        pass
    else:
        produto_id = request.GET.get('option', None)

        if produto_id == '':
            produto_id = '0'

        out = []

        try:
            lotes = EntradaProduto.objects.filter(produto_id=produto_id).order_by('data_validade')
        except EntradaProduto.DoesNotExist:
            lotes = None

        if lotes:
            out.append(u'{"id":0, "nome":"--- Selecione ---"}')
            for l in lotes:
                out.append(u'{"id":%s, "nome":"%s"}' % (l.id, l.lote))

        return HttpResponse(json.dumps(out), content_type='application/javascript')


@login_required(login_url='/login')
def get_qtdelote(request):
    if request.method != 'GET' or request.is_ajax() is False:
        pass
    else:
        entrada_id = request.GET.get('option', None)

        if entrada_id == '':
            entrada_id = '0'

        out = []

        try:
            lote = EntradaProduto.objects.get(id=entrada_id)
        except EntradaProduto.DoesNotExist:
            lote = None

        if lote:
            qtde_lote = lote.quantidade
            out.append(u'{"qtde_lote":%s}' % qtde_lote)

        return HttpResponse(json.dumps(out), content_type='application/javascript')


@login_required(login_url='/login')
def lista_movimento(request):
    primeiro_dia = request.GET.get("data_inicial")
    ultimo_dia = request.GET.get("data_final")
    ordenacao = request.GET.get("ordenacao", "-data_cadastro")

    parametros_filtro = {}
    data_inicial = None
    data_final = None

    if primeiro_dia and ultimo_dia:
        data_inicial = timezone.datetime.strptime(request.GET.get("data_inicial", primeiro_dia) + " 00:00",
                                                  "%Y-%m-%d %H:%M")
        data_final = timezone.datetime.strptime(request.GET.get("data_final", ultimo_dia) + " 23:59:59",
                                                "%Y-%m-%d %H:%M:%S")
        intervalo_datas = [data_inicial, data_final]
        parametros_filtro.update({
            "data_mov__range": intervalo_datas,
        })

    str_busca = request.GET.urlencode(True)
    if len(str_busca) <= 7:
        str_busca = ''

    lista_mov_entrada = EntradaProduto.objects.annotate(
        descricao_mov=F('descricao'),
        lote_mov=F('lote'),
        nome_produto=F('produto__nome'),
        nome_categoria=F('produto__categoria__nome'),
        fornecedor_mov=F('fornecedor__nome_fantasia'),
        usuario_mov=F('usuario__username'),
        quantidade_mov=F('quantidade_inicial'),
        data_mov=F('data_cadastro')).extra(select={'tipo_mov': 1}).values(
        'descricao_mov',
        'lote_mov',
        'nome_produto',
        'nome_categoria',
        'fornecedor_mov',
        'usuario_mov',
        'quantidade_mov',
        'data_mov',
        'tipo_mov').filter(**parametros_filtro)

    lista_mov_saida = SaidaProduto.objects.annotate(
        descricao_mov=F('descricao'),
        lote_mov=F('entrada__lote'),
        nome_produto=F('entrada__produto__nome'),
        nome_categoria=F('entrada__produto__categoria__nome'),
        fornecedor_mov=F('entrada__fornecedor__nome_fantasia'),
        usuario_mov=F('usuario__username'),
        quantidade_mov=F('quantidade'),
        data_mov=F('data_cadastro')).extra(select={'tipo_mov': 2}).values(
        'descricao_mov',
        'lote_mov',
        'nome_produto',
        'nome_categoria',
        'fornecedor_mov',
        'usuario_mov',
        'quantidade_mov',
        'data_mov',
        'tipo_mov').filter(**parametros_filtro)

    # percorrer lista_mov_entrada e lista_mov_saida e juntar tudo em uma: lista_movimento

    lista_movimento = []

    for i_entrada in lista_mov_entrada:
        lista_movimento.append(i_entrada)

    for i_saida in lista_mov_saida:
        lista_movimento.append(i_saida)

    lista_movimento = sorted(lista_movimento, key=lambda k: k['data_mov'], reverse=ordenacao == '-data_cadastro')

    # -------------------- PAGINATOR ------------------------

    paginator = Paginator(lista_movimento, 9)
    page = request.GET.get('page')
    try:
        objs = paginator.page(page)
    except PageNotAnInteger:
        objs = paginator.page(1)
    except EmptyPage:
        objs = paginator.page(paginator.num_pages)

    query_string = u""
    if request.GET:
        for i in request.GET:
            if i != "page":
                for v in request.GET.getlist(i):
                    query_string += u"&%s=%s" % (i, quote(v))

    # ------------------- FIM PAGINATOR ----------------------

    # --------------------- PDF ------------------------------

    if request.GET.get("export") == "pdf":
        data = datetime.date.today()
        return render_to_pdf("relatorios/movimentos.html", locals(),
                             "movimentos-%s.pdf" % data.isoformat())

    # --------------------- FIM PDF ---------------------------

    # ------------------------ XLS ----------------------------

    if request.GET.get("export") == "xls":
        data = datetime.date.today()
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="movimentos-%s.xlsx' % data.isoformat()
        wb = Workbook()
        ws = wb.get_active_sheet()

        row_num = 1
        titulo = u"Relatório de Movimentos"

        if data_inicial and data_final:
            titulo += u" - Referência %s até %s" % (data_inicial, data_final)

        c = ws.cell(row=row_num, column=1)
        c.value = titulo
        columns = [
            (u"Descrição", 70),
            (u"Lote", 70),
            (u"Produto", 70),
            (u"Categoria", 70),
            (u"Fornecedor", 70),
            (u"Usuário", 70),
            (u"Qtde", 70),
            (u"Dt.Movimento", 70),
            (u"Tipo", 70),
        ]

        for col_num in range(len(columns)):
            c = ws.cell(row=row_num + 1, column=col_num + 1)
            c.value = columns[col_num][0]

        for obj in lista_movimento:
            row_num += 1

            tipo_movimento = 'Entrada'

            if obj.get('tipo_mov') == 2:
                tipo_movimento = 'Saída'

            row = [
                obj.get('descricao_mov'),
                obj.get('lote_mov'),
                obj.get('nome_produto'),
                obj.get('nome_categoria'),
                obj.get('fornecedor_mov'),
                obj.get('usuario_mov'),
                obj.get('quantidade_mov'),
                "{0:%d/%m/%Y} ás {0:%H:%M:%S}".format(obj.get('data_mov')),
                tipo_movimento
            ]
            for col_num in range(len(row)):
                c = ws.cell(row=row_num + 1, column=col_num + 1)
                c.value = u"%s" % row[col_num]

        rodape = u"Relatório emitido por {0} em {1:%d/%m/%Y %H:%M:%S}".format(request.user.username,
                                                                              datetime.datetime.now())
        c = ws.cell(row=row_num + 3, column=1)
        c.value = rodape
        wb.save(response)
        return response

    # -----------------------FIM XLS---------------------------

    return render(request, 'principal/lista_movimento.html', locals())
